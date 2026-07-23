"""
Azure Cost Dashboard Builder
=============================
Reusable tool: reads every Azure cost export (.xlsx/.csv) in the data/ folder,
combines them, and rebuilds an Excel dashboard with:
  - Monthly cost consumption trend
  - Month-on-month % increase/decrease
  - Cost breakdown by service

Re-run any time you drop a new monthly export into data/:
    python analyze_azure_costs.py

Column detection is flexible: it looks for common Azure Cost Management export
column names (case/spacing-insensitive) for date, cost, service, resource group
and subscription. If your export uses different headers, add them to the
*_CANDIDATES lists below.
"""
import glob
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")
OUTPUT_FILE = os.path.join(HERE, "Azure_Cost_Dashboard.xlsx")

DATE_CANDIDATES = ["usagedatetime", "usagedate", "date", "billingperiodstartdate", "billingmonth"]
COST_CANDIDATES = ["costinbillingcurrency", "pretaxcost", "cost", "costusd", "billedcost", "effectivecost"]
SERVICE_CANDIDATES = ["servicename", "metercategory", "consumedservice"]
SUB_CANDIDATES = ["subscriptionname", "subscription", "subscriptionid"]
RG_CANDIDATES = ["resourcegroupname", "resourcegroup"]

FONT_NAME = "Arial"
NAVY = "1F3864"
LIGHT_BLUE = "D9E2F3"
WHITE_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="B7B7B7")] * 4)
CURRENCY_FMT = '$#,##0.00;($#,##0.00);"-"'
PERCENT_FMT = "0.0%;-0.0%;\"-\""


def find_col(columns, candidates):
    norm = {str(c).lower().replace(" ", "").replace("_", ""): c for c in columns}
    for cand in candidates:
        if cand in norm:
            return norm[cand]
    return None


def load_all_files(data_dir):
    files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")) + glob.glob(os.path.join(data_dir, "*.csv")))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    frames = []
    for f in files:
        try:
            df = pd.read_csv(f) if f.lower().endswith(".csv") else pd.read_excel(f)
        except Exception as exc:
            print(f"  Skipping {os.path.basename(f)}: could not read file ({exc})")
            continue

        date_col = find_col(df.columns, DATE_CANDIDATES)
        cost_col = find_col(df.columns, COST_CANDIDATES)
        if not date_col or not cost_col:
            print(f"  Skipping {os.path.basename(f)}: no recognizable date/cost column (found: {list(df.columns)})")
            continue

        service_col = find_col(df.columns, SERVICE_CANDIDATES)
        sub_col = find_col(df.columns, SUB_CANDIDATES)
        rg_col = find_col(df.columns, RG_CANDIDATES)

        out = pd.DataFrame()
        out["Date"] = pd.to_datetime(df[date_col], errors="coerce")
        out["Cost"] = pd.to_numeric(df[cost_col], errors="coerce")
        out["Service"] = df[service_col].astype(str) if service_col else "Unspecified"
        out["Subscription"] = df[sub_col].astype(str) if sub_col else "Unspecified"
        out["ResourceGroup"] = df[rg_col].astype(str) if rg_col else "Unspecified"
        out["SourceFile"] = os.path.basename(f)
        out = out.dropna(subset=["Date", "Cost"])
        print(f"  Loaded {os.path.basename(f)}: {len(out)} rows, ${out['Cost'].sum():,.2f}")
        frames.append(out)

    if not frames:
        sys.exit(f"No usable Azure cost files found in {data_dir}. Add .xlsx/.csv exports and re-run.")
    return pd.concat(frames, ignore_index=True)


def style_header_row(ws, row, ncols, fill=HEADER_FILL, font=WHITE_FONT):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def autosize(ws, ncols, min_width=12, max_width=40):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = max(
            (len(str(ws.cell(row=r, column=col).value)) for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=col).value is not None),
            default=min_width,
        )
        ws.column_dimensions[letter].width = min(max(longest + 2, min_width), max_width)


def build_workbook(combined):
    combined = combined.sort_values("Date")
    months = sorted(combined["Date"].dt.to_period("M").astype(str).unique())
    services = sorted(combined["Service"].unique())

    wb = Workbook()
    wb.remove(wb.active)

    # ---------- Raw Data ----------
    ws_raw = wb.create_sheet("Raw Data")
    headers = ["Date", "Month", "Cost", "Service", "Resource Group", "Subscription", "Source File"]
    ws_raw.append(headers)
    style_header_row(ws_raw, 1, len(headers))
    for i, r in enumerate(combined.itertuples(index=False), start=2):
        ws_raw.cell(row=i, column=1, value=r.Date.to_pydatetime()).number_format = "yyyy-mm-dd"
        ws_raw.cell(row=i, column=2, value=f'=TEXT(A{i},"yyyy-mm")')
        ws_raw.cell(row=i, column=3, value=round(float(r.Cost), 2)).number_format = CURRENCY_FMT
        ws_raw.cell(row=i, column=4, value=r.Service)
        ws_raw.cell(row=i, column=5, value=r.ResourceGroup)
        ws_raw.cell(row=i, column=6, value=r.Subscription)
        ws_raw.cell(row=i, column=7, value=r.SourceFile)
    last_raw_row = ws_raw.max_row
    ws_raw.freeze_panes = "A2"
    autosize(ws_raw, len(headers))
    for font_name_row in range(1, last_raw_row + 1):
        for col in range(1, len(headers) + 1):
            c = ws_raw.cell(row=font_name_row, column=col)
            if font_name_row > 1:
                c.font = Font(name=FONT_NAME)

    # ---------- Monthly Summary ----------
    ws_sum = wb.create_sheet("Monthly Summary")
    headers = ["Month", "Month Label", "Total Cost", "MoM Change ($)", "MoM Change (%)"]
    ws_sum.append(headers)
    style_header_row(ws_sum, 1, len(headers))
    for i, month in enumerate(months, start=2):
        ws_sum.cell(row=i, column=1, value=month)
        ws_sum.cell(row=i, column=2, value=f'=TEXT(DATEVALUE(A{i}&"-01"),"mmm yyyy")')
        ws_sum.cell(row=i, column=3, value=f"=SUMIFS('Raw Data'!C:C,'Raw Data'!B:B,A{i})").number_format = CURRENCY_FMT
        if i == 2:
            ws_sum.cell(row=i, column=4, value='="-"')
            ws_sum.cell(row=i, column=5, value='="-"')
        else:
            ws_sum.cell(row=i, column=4, value=f"=C{i}-C{i - 1}").number_format = CURRENCY_FMT
            ws_sum.cell(row=i, column=5, value=f"=IFERROR((C{i}-C{i - 1})/C{i - 1},\"-\")").number_format = PERCENT_FMT
    last_sum_row = ws_sum.max_row
    for r in range(2, last_sum_row + 1):
        for col in range(1, len(headers) + 1):
            ws_sum.cell(row=r, column=col).font = Font(name=FONT_NAME)
            ws_sum.cell(row=r, column=col).border = THIN_BORDER
    ws_sum.freeze_panes = "A2"
    autosize(ws_sum, len(headers))

    # ---------- By Service ----------
    ws_svc = wb.create_sheet("By Service")
    headers = ["Month"] + services + ["Total"]
    ws_svc.append(headers)
    style_header_row(ws_svc, 1, len(headers))
    for i, month in enumerate(months, start=2):
        ws_svc.cell(row=i, column=1, value=month)
        for j, svc in enumerate(services, start=2):
            col_letter = get_column_letter(j)
            formula = f"=SUMIFS('Raw Data'!C:C,'Raw Data'!B:B,$A{i},'Raw Data'!D:D,{col_letter}$1)"
            ws_svc.cell(row=i, column=j, value=formula).number_format = CURRENCY_FMT
        total_col = len(services) + 2
        first_letter = get_column_letter(2)
        last_letter = get_column_letter(total_col - 1)
        ws_svc.cell(row=i, column=total_col, value=f"=SUM({first_letter}{i}:{last_letter}{i})").number_format = CURRENCY_FMT
    last_svc_row = ws_svc.max_row
    for r in range(2, last_svc_row + 1):
        for col in range(1, len(headers) + 1):
            ws_svc.cell(row=r, column=col).font = Font(name=FONT_NAME)
    ws_svc.freeze_panes = "B2"
    autosize(ws_svc, len(headers))

    # ---------- Dashboard ----------
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False
    ws_dash["B2"] = "Azure Cost Dashboard"
    ws_dash["B2"].font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
    ws_dash["B3"] = f"Data through {months[-1]} | Source files refreshed on each run"
    ws_dash["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color="666666")

    kpis = [
        ("Latest Month", f"='Monthly Summary'!B{last_sum_row}", None),
        ("Latest Month Cost", f"='Monthly Summary'!C{last_sum_row}", CURRENCY_FMT),
        ("MoM Change (Latest)", f"='Monthly Summary'!E{last_sum_row}", PERCENT_FMT),
        ("Total Cost (All Months)", f"=SUM('Monthly Summary'!C2:C{last_sum_row})", CURRENCY_FMT),
        ("Average Monthly Cost", f"=AVERAGE('Monthly Summary'!C2:C{last_sum_row})", CURRENCY_FMT),
    ]
    kpi_start_row = 5
    kpi_col_width = 2
    for idx, (label, formula, fmt) in enumerate(kpis):
        col = 2 + idx * kpi_col_width
        letter = get_column_letter(col)
        ws_dash.merge_cells(f"{letter}{kpi_start_row}:{get_column_letter(col + 1)}{kpi_start_row}")
        lbl_cell = ws_dash[f"{letter}{kpi_start_row}"]
        lbl_cell.value = label
        lbl_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="FFFFFF")
        lbl_cell.fill = HEADER_FILL
        lbl_cell.alignment = Alignment(horizontal="center")
        ws_dash.merge_cells(f"{letter}{kpi_start_row + 1}:{get_column_letter(col + 1)}{kpi_start_row + 1}")
        val_cell = ws_dash[f"{letter}{kpi_start_row + 1}"]
        val_cell.value = formula
        if fmt:
            val_cell.number_format = fmt
        val_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=NAVY)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = SUBHEADER_FILL

    # Line chart: monthly cost trend
    trend_chart = LineChart()
    trend_chart.title = "Monthly Cost Consumption Trend"
    trend_chart.style = 10
    trend_chart.y_axis.title = "Cost (USD)"
    trend_chart.x_axis.title = "Month"
    trend_chart.height = 9
    trend_chart.width = 22
    data_ref = Reference(ws_sum, min_col=3, min_row=1, max_row=last_sum_row)
    cats_ref = Reference(ws_sum, min_col=2, min_row=2, max_row=last_sum_row)
    trend_chart.add_data(data_ref, titles_from_data=True)
    trend_chart.set_categories(cats_ref)
    trend_chart.series[0].smooth = False
    trend_chart.series[0].marker.symbol = "circle"
    ws_dash.add_chart(trend_chart, "B8")

    # Bar chart: MoM % change (skip first month, which has no prior month)
    mom_chart = BarChart()
    mom_chart.type = "col"
    mom_chart.title = "Month-on-Month Cost Change (%)"
    mom_chart.style = 11
    mom_chart.y_axis.title = "% Change vs Prior Month"
    mom_chart.x_axis.title = "Month"
    mom_chart.height = 9
    mom_chart.width = 22
    if last_sum_row >= 3:
        mom_data_ref = Reference(ws_sum, min_col=5, min_row=1, max_row=last_sum_row)
        mom_cats_ref = Reference(ws_sum, min_col=2, min_row=3, max_row=last_sum_row)
        mom_chart.add_data(mom_data_ref, titles_from_data=True)
        mom_chart.set_categories(mom_cats_ref)
        mom_chart.dLbls = DataLabelList()
        mom_chart.dLbls.showVal = True
    ws_dash.add_chart(mom_chart, "L8")

    # Stacked bar chart: cost by service
    svc_chart = BarChart()
    svc_chart.type = "col"
    svc_chart.grouping = "stacked"
    svc_chart.overlap = 100
    svc_chart.title = "Monthly Cost by Service"
    svc_chart.style = 12
    svc_chart.y_axis.title = "Cost (USD)"
    svc_chart.x_axis.title = "Month"
    svc_chart.height = 9
    svc_chart.width = 22
    svc_data_ref = Reference(ws_svc, min_col=2, max_col=1 + len(services), min_row=1, max_row=last_svc_row)
    svc_cats_ref = Reference(ws_svc, min_col=1, min_row=2, max_row=last_svc_row)
    svc_chart.add_data(svc_data_ref, titles_from_data=True)
    svc_chart.set_categories(svc_cats_ref)
    ws_dash.add_chart(svc_chart, "B24")

    ws_dash.column_dimensions["A"].width = 2
    for col in range(2, 20):
        ws_dash.column_dimensions[get_column_letter(col)].width = 11

    wb.save(OUTPUT_FILE)


def main():
    print(f"Reading Azure cost exports from: {DATA_DIR}")
    combined = load_all_files(DATA_DIR)
    print(f"\nCombined dataset: {len(combined)} rows across {combined['Date'].dt.to_period('M').nunique()} month(s)")
    build_workbook(combined)
    print(f"\nDashboard written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
